import streamlit as st
import pandas as pd
import io
import json
import os
import pdfkit  # Certifique-se de que o pdfkit está instalado
from openpyxl import Workbook
import tempfile
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi

def app():     


    path_to_wkhtmltopdf = '/bin/wkhtmltopdf'  
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)


    # Função para carregar arquivos JSON
    def carregar_json(arquivo):
        if os.path.exists(arquivo):
            with open(arquivo, 'r', encoding='utf-8') as json_file:
                return json.load(json_file)
        return None


    def conectar_base():
        if "mongo_client" not in st.session_state:
            uri = "mongodb+srv://jmfnet2004:xWYCXScI3Nfa4VJn@cluster0.cl7c2.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
            client = MongoClient(uri, server_api=ServerApi('1'))
            st.session_state.mongo_client = client
            st.session_state.db = st.session_state.mongo_client["ipog"]
            st.session_state.connection_initialized = True

    def fechar_base():
        if "mongo_client" in st.session_state:
            st.session_state.mongo_client.close()
            del st.session_state.mongo_client
            st.session_state.connection_initialized = False

    def fetch_collection_data(db_name, collection_name):
        db = st.session_state.db
        collection = db[collection_name]
        data = list(collection.find())
        for record in data:
            record.pop('_id', None)
        return data

    def clear_collection(db_name,collection_name):
        db = st.session_state.db
        collection = db[collection_name]
        result = collection.delete_many({})

    def insert_data(data, db_name, collection_name):
        db = st.session_state.db
        collection = db[collection_name]
        collection.insert_one(data)
       
    def carrega_tabelas():
        periodos_df = pd.DataFrame(fetch_collection_data("ipog", "periodos"))
        professores = pd.DataFrame(fetch_collection_data("ipog", "professores"))
        cursos = pd.DataFrame(fetch_collection_data("ipog", "cursos"))
        modalidades = pd.DataFrame(fetch_collection_data("ipog", "modalidades"))

        horarios = fetch_collection_data("ipog", "horarios")[0]['horarios']
        dias = fetch_collection_data("ipog", "dias")[0]['dias']

        cursos = cursos.sort_values(by="nome_curso", ascending=True)
        professores = professores.sort_values(by="nome_professor", ascending=True)
        modalidades = modalidades.sort_values(by="nome_modalidade", ascending=True)
        periodos_df = periodos_df.sort_values(by="nome_periodo", ascending=True)


        return professores, cursos, periodos_df, horarios, dias, modalidades

    def carregar_disciplinas(course_name):
        colecao = f'disciplinas_{course_name.replace(" ", "_").lower()}'

        disciplinas =  pd.DataFrame(fetch_collection_data("ipog", colecao)[0]['disciplinas'])
        disciplinas = disciplinas.sort_values(by="nome_disciplina", ascending=True)
        return disciplinas

    def carregar_alocacao(course_name):
        
        colecao = f'alocacoes_{course_name.replace(" ", "_").lower()}'
        dados = fetch_collection_data("ipog", colecao)
        return dados

    def carrega_dados_tabelas():
        disciplinas = carregar_disciplinas(curso_selecionado)

        # Filtrar as disciplinas do curso selecionado
        disciplinas_opcoes = disciplinas[disciplinas['id_curso'] == id_curso_selecionado]['nome_disciplina'].tolist()
        professores_opcoes = professores.apply(lambda row: f"{row['matricula']} - {row['nome_professor']}", axis=1).tolist()
        modalidades_opcoes = modalidades['nome_modalidade'].tolist()
        return disciplinas, disciplinas_opcoes, professores_opcoes, modalidades_opcoes

    def carregar_dados_na_session_state(dados_json_, course_id):
        if dados_json_:
            try:
                dados_json = dados_json_[0]
            except:
                dados_json = dados_json_

            for periodo, alocacoes in dados_json["periodos"].items():
                if course_id not in st.session_state['agendas']:
                    st.session_state['agendas'][course_id] = {}
        
                if periodo not in st.session_state['registros'][course_id]:
                    st.session_state['registros'][course_id][periodo] = create_schedule_table()
                    st.session_state['agendas'][course_id][periodo] = create_schedule_table()

                for alocacao in alocacoes:
                    disciplina = alocacao["disciplina"]
                    matricula_professor = alocacao["matricula_professor"]
                    professor = alocacao["professor"]
                    horario = alocacao["horario"]
                    dia = alocacao["dia"]
                    modalidade = alocacao["modalidade"]
                
                    st.session_state['registros'][course_id][periodo].loc[horario, dia] = f"{disciplina} / {matricula_professor} / {professor} / {modalidade}"
                    st.session_state['agendas'][course_id][periodo].loc[horario, dia] = f"{disciplina} / {matricula_professor} / {professor} / {modalidade}"
                 
    def carrega_planilha_horario(id_curso_selecionado):

        df = pd.read_excel(uploaded_file, sheet_name="Dados")
        df['Curso'] = df['Curso'].ffill()
        df['Horário'] = df['Horário'].ffill()
        df['Período'] = df['Período'].ffill()
        df['Horário'] = df['Horário'].str.replace('–','-', regex=False)

        # Obter o nome do curso
        curso_nome = df['Curso'].iloc[0]

        # Inicializar estrutura de saída com o nome do curso
        output = {
            "curso": curso_nome,
            "periodos": {}
        }

        # Iterar sobre cada período, horário e dia da semana
        for period, period_df in df.groupby("Período"):
            period_list = []
            for horario, horario_df in period_df.groupby("Horário"):
                for day in ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]:
                    day_data = horario_df[day].dropna()
                    if not day_data.empty:
                        item = {
                            "horario": horario,
                            "dia": day,
                        }
                        for _, row in horario_df.iterrows():
                            info_key = row["Informação"].lower()
                            item[info_key] = row[day]

                        period_list.append(item)

            if period_list:
                output["periodos"][period] = period_list


        st.session_state['registros'][id_curso_selecionado] = {}
        st.session_state['decisoes'] = {}

        incializa_agenda(id_curso_selecionado) 
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}
        
        carregar_dados_na_session_state(output, id_curso_selecionado)

        registrar_horario()
   
    def create_schedule_table():
        schedule = pd.DataFrame("", index=horarios, columns=dias)
        return schedule
    
    def allocate_schedule(course_id, period, day, time, discipline, professor, modalidade):
        st.session_state['agendas'][course_id][period].loc[time, day] = f"{discipline} / {professor} / {modalidade}"

    def deallocate_schedule(course_id, period, day, time):
        st.session_state['agendas'][course_id][period].loc[time, day] = ""

    def preenche_alocacao(id_curso_selecionado):
        
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}
            dados_json = carregar_alocacao(curso_selecionado)

            carregar_dados_na_session_state(dados_json, id_curso_selecionado)

    def salvar_horario(course_id, course_name):
        registros = st.session_state['registros'][course_id]
        data_para_salvar = {
            "curso": course_name,
            "periodos": {}
        }
        for periodo, tabela in registros.items():
            data_para_salvar["periodos"][periodo] = []
            for time in tabela.index:
                for day in tabela.columns:
                    cell_content = tabela.loc[time, day]
                    if cell_content:
                        disciplina, matricula_professor, professor, modalidade = cell_content.split(' / ')
                        data_para_salvar["periodos"][periodo].append({
                            "disciplina": disciplina,
                            "matricula_professor": matricula_professor,
                            "professor": professor,
                            "horario": time,
                            "dia": day,
                            "modalidade": modalidade
                        })
        
        colecao = f'alocacoes_{course_name.replace(" ", "_").lower()}'
        clear_collection('ipog',colecao)
        insert_data(data_para_salvar, 'ipog', colecao)

    def registrar_horario():
        # Inicializar 'registros' no session_state para o curso selecionado, se não existir
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}

        # Registrar o horário do período selecionado
        st.session_state['registros'][id_curso_selecionado][periodo_selecionado] = st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].copy()
        
        salvar_horario(id_curso_selecionado, curso_selecionado)
       
    def limpar_secao_ao_trocar_curso():
        st.session_state['agendas'] = {}
        st.session_state['registros'] = {}
        st.session_state['decisoes'] = {}
        #st.session_state['conflitos'] = {}
    
    def gerar_excel():
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            start_row = 0
            curso_nome = cursos[cursos['id_curso'] == id_curso_selecionado]['nome_curso'].values[0]
            writer.sheets['Horário'] = writer.book.create_sheet("Horário")
            worksheet = writer.sheets['Horário']
            worksheet.cell(row=1, column=1, value=f"Curso: {curso_nome}")
            start_row += 2
            for periodo, tabela in st.session_state['registros'][id_curso_selecionado].items():
                worksheet.cell(row=start_row, column=1, value=f"Período: {periodo}")
                start_row += 1
                tabela.to_excel(writer, sheet_name="Horário", startrow=start_row, index=True)
                start_row += len(tabela) + 3
        output.seek(0)
        return output   

    def limpar_horario():
        st.session_state['agendas'] = {}
        st.session_state['registros'] = {}
        st.session_state['decisoes'] = {}

        # Remover o arquivo JSON específico do curso selecionado
        
        course_name = st.session_state['curso_selecionado']
        db_name =  st.session_state.db
        colecao = f'alocacoes_{course_name.replace(" ", "_").lower()}'
        clear_collection(db_name,colecao)

        #course_name = st.session_state['curso_selecionado']
        #file_path = f'horarios/{course_name}_alocacoes.json'
        #file_path2 = f'juncoes/{course_name}_juncao.json'

        #if os.path.exists(file_path2):
        #    os.remove(file_path2)

        #if os.path.exists(file_path):
        #    os.remove(file_path)
        #    st.warning(f"Horário do curso {course_name} removido.")

    def reset_confirmacao_exclusao():
        # Inicializar a flag no session_state se ainda não existir
        if 'reset_confirmacao_exclusao' not in st.session_state:
            st.session_state['reset_confirmacao_exclusao'] = False

        # Verificar se precisamos redefinir o checkbox
        if st.session_state['reset_confirmacao_exclusao']:
            st.session_state['confirmacao_exclusao'] = False
            st.session_state['reset_confirmacao_exclusao'] = False  

        # Criar a checkbox com a chave especificada
        st.checkbox(
            "Desejo limpar o horário e excluir o arquivo de alocação desse curso. CUIDADO: Não tem como recuperar o Horário excluído.",
            key='confirmacao_exclusao'
        )
        
        
            # Adicionar estilo CSS para as colunas da tabela
    
    def reset_confirmacao_upload():
        # Inicializar a flag no session_state se ainda não existir
        if 'reset_confirmacao_upload' not in st.session_state:
            st.session_state['reset_confirmacao_upload'] = False

        # Verificar se precisamos redefinir o checkbox
        if st.session_state['reset_confirmacao_upload']:
            st.session_state['confirmacao_upload'] = False
            st.session_state['reset_confirmacao_upload'] = False

        # Checkbox de confirmação
        st.checkbox(
            "Desejo carregar uma nova planilha de horário que substituirá o atual (CUIDADO: não tem como recuperar o horário antigo.)",
            key='confirmacao_upload'
        )

    def adicionar_css_customizado():
        st.markdown(
            """
            <style>
            .schedule-table {
                table-layout: fixed;
                width: 100%;
                word-wrap: break-word;
                border-collapse: collapse;
                font-size: 10px; 
            }
            .schedule-table th, .schedule-table td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
                vertical-align: top;
                width: 100px;
                white-space: pre-wrap;
                word-wrap: break-word;
            }
            .schedule-table th {
                background-color: #f2f2f2;
            }
            
            .schedule {
                background-color: #f2f2f2;
                font-size: 10px; 
                border-radius: 10px;
                text-align: center;
            }
            .conflito {
                background-color: #F08080;
                color: black;
            }
            .fusao {
                background-color: #87CEEB;
                color: black;
            }
    

            .custom-button {
                font-size: 16px; /* Tamanho do texto */
                padding: 8px 16px; /* Ajusta o tamanho do botão */
                background-color: #ff4d4d; /* Cor do botão */
                color: white; /* Cor do texto */
                border: none;
                border-radius: 5px;
                cursor: pointer;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            """,
            unsafe_allow_html=True
        )

    def inicializa_sessao():
        # Inicializar as agendas por curso e período no session_state
        if 'agendas' not in st.session_state:
            st.session_state['agendas'] = {}

        # Inicializar registro de tabelas para períodos
        if 'registros' not in st.session_state:
            st.session_state['registros'] = {}

        # Inicializar 'decisoes' no session_state se ainda não estiver presente
        if 'decisoes' not in st.session_state:
            st.session_state['decisoes'] = {}

        if 'conflitos' not in st.session_state:
            st.session_state['conflitos'] = {}

    def incializa_agenda(id_curso_selecionado):
        # Inicializar a agenda para o curso e período selecionado, se não existir
        if id_curso_selecionado not in st.session_state['agendas']:
            st.session_state['agendas'][id_curso_selecionado] = {}

        if periodo_selecionado not in st.session_state['agendas'][id_curso_selecionado]:
            st.session_state['agendas'][id_curso_selecionado][periodo_selecionado] = create_schedule_table()

    def exibir_tabelas_periodo():
        st.write("### Horários Definidos por Período")
        if id_curso_selecionado in st.session_state['registros']:
            for periodo in periodos_df['nome_periodo']:
                if periodo in st.session_state['registros'][id_curso_selecionado]:
                    tabela = st.session_state['registros'][id_curso_selecionado][periodo].copy()

                    # Filtrar para exibir apenas "disciplina", "professor" e "modalidade" em cada célula
                    for time in tabela.index:
                        for day in tabela.columns:
                            cell_content = tabela.loc[time, day]
                            if pd.notna(cell_content) and cell_content:
                                # Extrair apenas disciplina e nome do professor
                                parts = cell_content.split(" / ")
                                if len(parts) >= 2:
                                    disciplina = parts[0]
                                    professor = parts[2]
                                    modalidade = parts[3]
                                    tabela.loc[time, day] = f"{disciplina} </br> {professor} </br> {modalidade}"

                    # Exibir a tabela estilizada
                    st.write(f"**Período: {periodo}**")
                    #st.write(tabela)
                    #st.data_editor(tabela)
                    #st.dataframe(tabela,
                    #        column_config={
                    #            "widgets": st.column_config.Column(
                    #                width="small"
                    #            )
                    #        })
                    st.write(tabela.to_html(escape=False, classes="schedule-table"), unsafe_allow_html=True)

    def renderizar_agenda(id_curso_selecionado, periodo_selecionado):
        cols = st.columns(len(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns) + 1)
        cols[0].write("**Horários**")
        for i, day in enumerate(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns):
            cols[i + 1].write(f"**{day}**")

        for time_slot in horarios:
            cols = st.columns(len(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns) + 1)
            cols[0].write(f"**{time_slot}**")
            for i, day in enumerate(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns):
                cell_content = st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].loc[time_slot, day]
                if cell_content:
                    
                    parts = cell_content.split(" / ")
                    disciplina = parts[0]
                    matricula = parts[1]
                    professor = parts[2]
                    modalidade = parts[3]
                    data = f"{disciplina} </br> {professor} ({matricula}) </br> {modalidade}"
                    
                    cols[i + 1].markdown(f"<div class='schedule'>{data}</div>", unsafe_allow_html=True)
                    if cols[i + 1].button("🗑️",  key=f"desalocar_{time_slot}_{day}"):
                        deallocate_schedule(id_curso_selecionado, periodo_selecionado, day, time_slot)
                        st.rerun()
                else:
                    cols[i + 1].markdown("<div class='schedule'>-</div>", unsafe_allow_html=True)

    def gerar_pdf_tabelas_periodo(curso_selecionado):

        # Criar um buffer de memória para o PDF
        pdf_buffer = io.BytesIO()
        
        # HTML inicial para o PDF
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif; 
                    font-size: 12px; 
                }}
                .schedule-table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                }}
                .schedule-table th, .schedule-table td {{ 
                    border: 1px solid #ddd; 
                    padding: 8px; 
                    text-align: center; 
                    width: 100px; 
                    word-wrap: break-word; 
                    white-space: pre-wrap; 
                }}
                .schedule-table th {{ background-color: #f2f2f2; }}
                .conflito {{ background-color: #F08080; color: black; }}
                .fusao {{ background-color: #87CEEB; color: black; }}
                 /* Estilização para as colunas de Sexta, Sábado e Domingo */
                .schedule-table th:nth-child(6), /* Sexta */
                .schedule-table th:nth-child(7), /* Sábado */
                .schedule-table th:nth-child(8), /* Domingo */
                .schedule-table td:nth-child(6),
                .schedule-table td:nth-child(7),
                .schedule-table td:nth-child(8) {{
                    background-color: #F5F5DC;
                }}
            </style>
        </head>
        <body>
        <h2>Horário de {curso_selecionado}</h2>
        """

        # Adicionar o conteúdo HTML de cada tabela de período
        if id_curso_selecionado in st.session_state['registros']:
            for periodo in periodos_df['nome_periodo']:
                if periodo in st.session_state['registros'][id_curso_selecionado]:
                    tabela = st.session_state['registros'][id_curso_selecionado][periodo].copy()

                    # Estilizar a tabela de acordo com a função `exibir_tabelas_periodo`
                    for time in tabela.index:
                        for day in tabela.columns:
                            cell_content = tabela.loc[time, day]
                            if pd.notna(cell_content) and cell_content:
                                parts = cell_content.split(" / ")
                                if len(parts) >= 2:
                                    disciplina = parts[0]
                                    professor = parts[2]
                                    modalidade = parts[3]
                                    tabela.loc[time, day] = f"{disciplina} </br> {professor} </br> {modalidade}"

                    # Adicionar a tabela ao HTML
                    html += f"<h3>Período: {periodo}</h3>"
                    html += tabela.to_html(escape=False, classes="schedule-table")

    
        options = {
            "orientation": "Landscape"
        }
        # Gerar o PDF usando pdfkit e o HTML construído
        with tempfile.NamedTemporaryFile(delete=True, suffix=".pdf") as tmp_pdf:
            pdfkit.from_string(html, tmp_pdf.name, options=options)
            tmp_pdf.seek(0)  # Retornar para o início do arquivo temporário
            pdf_buffer.write(tmp_pdf.read())  # Lê o conteúdo para o buffer

        pdf_buffer.seek(0)  # Volta para o início do buffer
        return pdf_buffer      

    def escolhe_curso_periodo(cursos, periodos_df):
        col1, col2 = st.columns(2)
        with col1:
            curso_selecionado = st.selectbox("Selecione o curso:", cursos['nome_curso'], key="curso_selecionado", on_change=limpar_secao_ao_trocar_curso)
            reset_confirmacao_upload()
            uploaded_file = st.file_uploader("Carregar Planilha de Horário", type="xlsx") if st.session_state.get('confirmacao_upload') else None
        
        with col2: periodo_selecionado = st.selectbox("Selecione o período:", periodos_df['nome_periodo'].tolist())
        
        id_curso_selecionado = cursos[cursos['nome_curso'] == curso_selecionado]['id_curso'].values[0]
        return id_curso_selecionado, curso_selecionado,  periodo_selecionado, uploaded_file

    def escolhe_dia_horario(dias, horarios):
        # Combobox para as escolhas de "Dia da Semana" e "Horário" 
        st.write("### Alocação de Disciplina e Professor")
        col3, col4 = st.columns(2)
        with col3: day = st.selectbox("Dia da Semana", dias)
        with col4: time = st.selectbox("Horário", horarios)
        return day, time

    def escolhe_disciplina_modalidade(disciplinas_opcoes, modalidades_opcoes):
        # Combobox para as escolhas de "Disciplina" e "Professor"
        col5, col6 = st.columns(2)
        with col5:  disciplina = st.selectbox("Disciplina", disciplinas_opcoes)
        with col6: modalidade = st.selectbox("Modalidade", modalidades_opcoes)
        return disciplina, modalidade
    
    def  escolhe_professor(professores_opcoes):
        # Alocar disciplina e professor# Criar uma nova linha para os dois botões
        col7, col8 = st.columns(2)
        with col7:
            professor = st.selectbox("Professor", professores_opcoes)
            alocar_button = st.button("Alocar Disciplina e Professor", key="alocar", use_container_width=True)
        return professor, alocar_button
    
    def escolhe_registrar_horario(id_curso_selecionado, periodo_selecionado):
        col_button, _ = st.columns([1, 2])  
        with col_button:
            registrar_button = st.button("Salvar Período no Horário", key="registrar", use_container_width=True)         
        return registrar_button

    def escolhe_exclusao():
        reset_confirmacao_exclusao()   
        if st.session_state.get('confirmacao_exclusao'):
            if st.button("Limpar Horário e Recomeçar"):
                limpar_horario()
                st.success("Horário e arquivo de alocação excluídos com sucesso.")
                st.session_state['reset_confirmacao_exclusao'] = True 
                st.rerun()
        else:
            st.warning("Por favor, marque a confirmação para visualizar o botão de exclusão.")

    def escolhe_baixar_pdf(curso_selecionado):
        if st.button("Baixar PDF do Horário"):
            pdf_file = gerar_pdf_tabelas_periodo(curso_selecionado)
            st.download_button(
                label="Download PDF",
                data=pdf_file,
                file_name="horarios_geral.pdf",
                mime="application/pdf"
            )

    def escolher_gerar_excel():
        st.write("\n")
        if st.button("Gerar Excel"):
            excel_file = gerar_excel()
            st.download_button(label="Download Excel", data=excel_file, file_name="tabelas_registradas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
         
#####################################################################################

    st.title("📆 :blue[Montagem de Horário por Curso]")
    st.divider()
    conectar_base()
    salva_conflito = False
    adicionar_css_customizado()
    
    professores, cursos, periodos_df, horarios, dias, modalidades = carrega_tabelas()
    inicializa_sessao()

    id_curso_selecionado, curso_selecionado, periodo_selecionado, uploaded_file = escolhe_curso_periodo(cursos, periodos_df)

    preenche_alocacao(id_curso_selecionado)    
    incializa_agenda(id_curso_selecionado)     

    disciplinas, disciplinas_opcoes, professores_opcoes, modalidades_opcoes = carrega_dados_tabelas()    

    if uploaded_file and st.session_state.get('confirmacao_upload'):
        carrega_planilha_horario(id_curso_selecionado)
        st.session_state['reset_confirmacao_upload'] = True     
        st.success("Planilha carregada com sucesso.")

    day, time = escolhe_dia_horario(dias, horarios)
    disciplina, modalidade = escolhe_disciplina_modalidade(disciplinas_opcoes, modalidades_opcoes)


    professor, alocar_button = escolhe_professor(professores_opcoes)
    if alocar_button:
        allocate_schedule(id_curso_selecionado, periodo_selecionado, day, time, disciplina, professor, modalidade)
    

    # Renderizar a agenda de definição de horário
    st.write(f"### Programação de Horários - {periodo_selecionado}")
    renderizar_agenda(id_curso_selecionado, periodo_selecionado)

    registrar_button = escolhe_registrar_horario(id_curso_selecionado, periodo_selecionado)
    if registrar_button:
        registrar_horario()

    exibir_tabelas_periodo()
    escolhe_baixar_pdf(curso_selecionado)
    escolhe_exclusao()
    escolher_gerar_excel()
    
