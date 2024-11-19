import streamlit as st
import pandas as pd
import io
import json
import os
import pdfkit  # Certifique-se de que o pdfkit está instalado
from openpyxl import Workbook
import tempfile

def app():
    
    # Função para carregar arquivos JSON
    def carregar_json(arquivo):
        if os.path.exists(arquivo):
            with open(arquivo, 'r', encoding='utf-8') as json_file:
                return json.load(json_file)
        return None

    # Função para carregar as disciplinas de acordo com o curso selecionado
    def carregar_disciplinas(course_name):
        arquivo = f'matrizes/disciplinas_{course_name.replace(" ", "_").lower()}.json'
        if not os.path.exists(arquivo):
            st.warning(f"Disciplinas do curso {course_name} não cadastradas.")
            st.stop()
        with open(arquivo, 'r', encoding='utf-8') as json_file:
            disciplinas_json = json.load(json_file)
        if disciplinas_json:
            return pd.DataFrame(disciplinas_json['disciplinas'])
        return pd.DataFrame()

    def carregar_juncoes(course_name):
        # Definir o caminho do arquivo JSON com base no nome do curso
        pasta_juncao = "juncoes"
        nome_arquivo = f"{course_name}_juncao.json"
        caminho_arquivo = os.path.join(pasta_juncao, nome_arquivo)

        # Verificar se o arquivo JSON existe
        if os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, "r", encoding='utf-8') as f:
                json_decisoes = json.load(f)

            # Inicializar 'decisoes' no session_state se ainda não estiver presente
            if 'decisoes' not in st.session_state:
                st.session_state['decisoes'] = {}

            # Popular o session_state com as decisões do JSON
            for conflito in json_decisoes:
                # Construir a chave única do conflito no formato 'Matrícula_Dia_Horário'
                conflito_key = f"{conflito['Matrícula']}_{conflito['Dia']}_{conflito['Horário']}"

                # Armazenar a estrutura detalhada do conflito no session_state
                st.session_state['decisoes'][conflito_key] = {
                    "Professor": conflito["Professor"],
                    "Matrícula": conflito["Matrícula"],
                    "Dia": conflito["Dia"],
                    "Horário": conflito["Horário"],
                    "Tipo": conflito["Tipo"],
                    "Cursos": conflito["Cursos"]
                }
        else:
            if 'decisoes' not in st.session_state:
                st.session_state['decisoes'] = {}

    # Função para criar a agenda de horários e alocar disciplinas e professores
    def create_schedule_table():
        schedule = pd.DataFrame("", index=horarios, columns=dias)
        return schedule

    # Função para preencher a alocação na tabela de horários
    def allocate_schedule(course_id, period, day, time, discipline, professor, modalidade):
        st.session_state['agendas'][course_id][period].loc[time, day] = f"{discipline} - {professor} - {modalidade}"

    # Função para desalocar uma disciplina/professor da agenda
    def deallocate_schedule(course_id, period, day, time):
        st.session_state['agendas'][course_id][period].loc[time, day] = ""

    # Função para verificar se há choque de horários para um mesmo professor (usando matrícula)
    def verificar_choque(course_id):
        conflitos = {}
        conflitos_detalhes = []
        for period, agenda in st.session_state['registros'][course_id].items():
            for time in horarios:
                for day in agenda.columns:
                    cell_content = agenda.loc[time, day]
                    if cell_content:
                        matricula_professor = cell_content.split(' - ')[1]
                        if (day, time) not in conflitos:
                            conflitos[(day, time)] = (matricula_professor, period)
                        else:
                            if conflitos[(day, time)][0] == matricula_professor:
                                conflitos_detalhes.append((day, time, matricula_professor, conflitos[(day, time)][1], period))
        return conflitos_detalhes

    # Função para carregar os dados do arquivo JSON, se existir
    def carregar_dados_alocacao(course_name):
        if os.path.exists(f'horarios/{course_name}_alocacoes.json'):
            with open(f'horarios/{course_name}_alocacoes.json', 'r', encoding='utf-8') as json_file:
                dados = json.load(json_file)
                return dados
        return None

    # Função para salvar os dados de alocação em um arquivo JSON com a estrutura especificada
    def salvar_json(course_id, course_name):
        registros = st.session_state['registros'][course_id]
        data_para_salvar = {
            "curso": course_name,
            "periodos": {}
        }
        for periodo, tabela in registros.items():
            data_para_salvar["periodos"][periodo] = []
            for time in tabela.index:
                for day in tabela.columns:
                    cell_content = tabela.loc[time, day]
                    if cell_content:
                        disciplina, matricula_professor, professor, modalidade = cell_content.split(' - ')
                        data_para_salvar["periodos"][periodo].append({
                            "disciplina": disciplina,
                            "matricula_professor": matricula_professor,
                            "professor": professor,
                            "horario": time,
                            "dia": day,
                            "modalidade": modalidade
                        })
        with open(f'horarios/{course_name}_alocacoes.json', 'w', encoding='utf-8') as json_file:
            json.dump(data_para_salvar, json_file, indent=4)

    # Função para carregar os dados do JSON no session_state
    def carregar_dados_no_session_state(dados_json, course_id):
        if dados_json and "periodos" in dados_json:
            for periodo, alocacoes in dados_json["periodos"].items():
                
                #st.write(st.session_state['registros'][course_id])
                if course_id not in st.session_state['agendas']:
                    st.session_state['agendas'][course_id] = {}
    
                if periodo not in st.session_state['registros'][course_id]:
                    st.session_state['registros'][course_id][periodo] = create_schedule_table()
                    st.session_state['agendas'][course_id][periodo] = create_schedule_table()

                for alocacao in alocacoes:
                    disciplina = alocacao["disciplina"]
                    matricula_professor = alocacao["matricula_professor"]
                    professor = alocacao["professor"]
                    horario = alocacao["horario"]
                    dia = alocacao["dia"]
                    modalidade = alocacao["modalidade"]
                    
                    st.session_state['registros'][course_id][periodo].loc[horario, dia] = f"{disciplina} - {matricula_professor} - {professor} - {modalidade}"
                    st.session_state['agendas'][course_id][periodo].loc[horario, dia] = f"{disciplina} - {matricula_professor} - {professor} - {modalidade}"

    # Função verificar_conflitos_entre_cursos permanece a mesma
    def verificar_conflitos_entre_cursos(diretorio_horarios):
        horarios_professores = {}

        for arquivo in os.listdir(diretorio_horarios):
            if arquivo.endswith('_alocacoes.json'):
                caminho_arquivo = os.path.join(diretorio_horarios, arquivo)
                
                with open(caminho_arquivo, 'r', encoding='utf-8') as json_file:
                    dados = json.load(json_file)
                    nome_curso = dados.get("curso", "Curso desconhecido")
                    
                    for periodo, alocacoes in dados["periodos"].items():
                        for alocacao in alocacoes:
                            dia = alocacao["dia"]
                            horario = alocacao["horario"]
                            matricula_professor = alocacao["matricula_professor"]
                            nome_professor = alocacao["professor"]

                            chave_horario = (matricula_professor, nome_professor, dia, horario)
                            
                            if chave_horario in horarios_professores:
                                horarios_professores[chave_horario].append((nome_curso, periodo))
                            else:
                                horarios_professores[chave_horario] = [(nome_curso, periodo)]
        
        conflitos = []
        for chave, cursos in horarios_professores.items():
            if len(cursos) > 1:
                matricula_professor, nome_professor, dia, horario = chave
                conflito_info = {
                    "matricula_professor": matricula_professor,
                    "nome_professor": nome_professor,
                    "dia": dia,
                    "horario": horario,
                    "cursos": cursos
                }
                conflitos.append(conflito_info)
        
        return conflitos

    def salvar_juncoes(curso_selecionado):

        # Criar a pasta 'juncoes' se ela não existir
        pasta_juncao = "juncoes"
        if not os.path.exists(pasta_juncao):
            os.makedirs(pasta_juncao)

        # Definir o nome do arquivo com o nome do curso
        nome_arquivo = f"{curso_selecionado}_juncao.json"
        caminho_arquivo = os.path.join(pasta_juncao, nome_arquivo)

        # Salvar o JSON em um arquivo
        with open(caminho_arquivo, "w", encoding='utf-8') as f:
            json.dump(list(st.session_state['decisoes'].values()), f, ensure_ascii=False, indent=4)

        st.rerun()

    def registrar_conflitos(conflitos):
        if 'decisoes' not in st.session_state:
            st.session_state['decisoes'] = {}

        for i, conflito in enumerate(conflitos):
                # Definir a chave única do conflito
                conflito_key = f"{conflito['matricula_professor']}_{conflito['dia']}_{conflito['horario']}"

                # Selecionar o valor inicial com base na sessão, ou "Conflito" por padrão
                tipo_conflito = st.session_state['decisoes'].get(conflito_key, {}).get("Tipo", "Conflito")
          
                # Atualizar ou adicionar a decisão completa no session_state
                st.session_state['decisoes'][conflito_key] = {
                    "Professor": conflito["nome_professor"],
                    "Matrícula": conflito["matricula_professor"],
                    "Dia": conflito["dia"],
                    "Horário": conflito["horario"],
                    "Tipo": tipo_conflito,
                    "Cursos": [{"Curso": curso, "Período": periodo} for curso, periodo in conflito["cursos"]]
                }      

    def exibir_conflitos(conflitos):
        if conflitos:
            st.markdown("**Conflitos de Horário Entre Cursos Detectados:**\n")
            
            # Inicializar 'decisoes' no session_state como um dicionário vazio, caso ainda não exista
            if 'decisoes' not in st.session_state:
                st.session_state['decisoes'] = {}

            for i, conflito in enumerate(conflitos):
                st.markdown(f"**Professor {conflito['nome_professor']} (Matrícula {conflito['matricula_professor']})** com conflito no dia **{conflito['dia']}** às **{conflito['horario']}**:")
                
                # Exibir detalhes dos cursos e períodos em conflito
                for curso, periodo in conflito["cursos"]:
                    st.markdown(f"- Curso: {curso}, Período: {periodo}")

                # Definir a chave única do conflito
                conflito_key = f"{conflito['matricula_professor']}_{conflito['dia']}_{conflito['horario']}"

                # Selecionar o valor inicial com base na sessão, ou "Conflito" por padrão
                tipo_conflito = st.session_state['decisoes'].get(conflito_key, {}).get("Tipo", "Conflito")

                # Opção para o usuário decidir entre Conflito ou Fusão
                escolha_usuario = st.radio(
                    f"Deseja tratar o conflito acima como uma 'Fusão' ou 'Conflito'?",
                    options=('Conflito', 'Fusão'),
                    index=0 if tipo_conflito == "Conflito" else 1,  # Define o índice baseado na escolha inicial
                    key=conflito_key  # Usa conflito_key para garantir a exclusividade
                )

                # Atualizar ou adicionar a decisão completa no session_state
                st.session_state['decisoes'][conflito_key] = {
                    "Professor": conflito["nome_professor"],
                    "Matrícula": conflito["matricula_professor"],
                    "Dia": conflito["dia"],
                    "Horário": conflito["horario"],
                    "Tipo": escolha_usuario,
                    "Cursos": [{"Curso": curso, "Período": periodo} for curso, periodo in conflito["cursos"]]
                }

            salvar_juncoes(curso_selecionado)
        else:
            course_name = st.session_state['curso_selecionado']
            file_path = f'juncoes/{course_name}_juncao.json'
            if os.path.exists(file_path):
                os.remove(file_path)

    # Função para verificar conflitos de horários do curso selecionado com outros cursos
    def verificar_conflitos_com_curso_selecionado(diretorio_horarios, curso_selecionado):
        horarios_professores = {}
        conflitos = []
        
        # Carregar o nome do curso selecionado
        arquivo_selecionado = f'horarios/{curso_selecionado}_alocacoes.json'
        
        for arquivo in os.listdir(diretorio_horarios):
            if arquivo.endswith('_alocacoes.json'):
                caminho_arquivo = os.path.join(diretorio_horarios, arquivo)
                
                with open(caminho_arquivo, 'r', encoding='utf-8') as json_file:
                    dados = json.load(json_file)
                    nome_curso = dados.get("curso", "Curso desconhecido")
                    
                    for periodo, alocacoes in dados["periodos"].items():
                        for alocacao in alocacoes:
                            dia = alocacao["dia"]
                            horario = alocacao["horario"]
                            matricula_professor = alocacao["matricula_professor"]
                            nome_professor = alocacao["professor"]
                            chave_horario = (matricula_professor, nome_professor, dia, horario)
                            
                            if chave_horario in horarios_professores:
                                # Verifica se já existe um conflito na lista de conflitos
                                conflito_existe = next(
                                    (conflito for conflito in conflitos 
                                    if conflito["matricula_professor"] == matricula_professor 
                                    and conflito["dia"] == dia 
                                    and conflito["horario"] == horario), 
                                    None
                                )
                                
                                # Adiciona o novo curso e período à lista de cursos
                                horarios_professores[chave_horario].append((nome_curso, periodo))
                                
                                if conflito_existe:
                                    # Atualiza o conflito existente com o novo curso e período
                                    conflito_existe["cursos"].append((nome_curso, periodo))
                                else:
                                    # Se o conflito ainda não existe, cria uma nova entrada de conflito
                                    conflito_info = {
                                        "matricula_professor": matricula_professor,
                                        "nome_professor": nome_professor,
                                        "dia": dia,
                                        "horario": horario,
                                        "cursos": [(curso, period) for curso, period in horarios_professores[chave_horario]]
                                    }
                                    conflitos.append(conflito_info)
                            else:
                                # Adiciona a primeira alocação para o horário do professor
                                horarios_professores[chave_horario] = [(nome_curso, periodo)]
        
        return conflitos

    # Registrar o período e verificar conflitos
    def registrar_horario():
        # Inicializar 'registros' no session_state para o curso selecionado, se não existir
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}

        # Registrar o horário do período selecionado
        st.session_state['registros'][id_curso_selecionado][periodo_selecionado] = st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].copy()
        
        # Verificar conflitos
        conflitos_detalhes = verificar_choque(id_curso_selecionado)
        salvar_json(id_curso_selecionado, curso_selecionado)
       

        if conflitos_detalhes:
            for i, conflito in enumerate(conflitos_detalhes):
                matricula_professor = conflito[2]
                nome_professor = professores[professores['matricula'] == int(matricula_professor)]['nome_professor'].values[0]
                conflito_key = f"{matricula_professor}_{conflito[0]}_{conflito[1]}_{i}"

                # Verificar se o conflito não está na sessão ou se a decisão é marcada como "Conflito"
                if conflito_key not in st.session_state['decisoes'] or st.session_state['decisoes'][conflito_key] == "Conflito":
                    st.error(
                        f"Conflito de horário: O(A) professor(a) {nome_professor} (Matrícula: {matricula_professor}) está alocado(a) "
                        f"no {conflito[3]} e {conflito[4]} para o mesmo horário ({conflito[0]} - {conflito[1]}) no curso {curso_selecionado}."
                    )
        else:
            st.success("Horário registrado com sucesso.")

    # Monitorar a seleção do curso para limpar a seção anterior
    def limpar_secao_ao_trocar_curso():
        st.session_state['agendas'] = {}
        st.session_state['registros'] = {}
        st.session_state['decisoes'] = {}
        st.session_state['conflitos'] = {}
    
    def gerar_excel():
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            start_row = 0
            curso_nome = cursos[cursos['id_curso'] == id_curso_selecionado]['nome_curso'].values[0]
            writer.sheets['Tabelas_Registradas'] = writer.book.create_sheet("Tabelas_Registradas")
            worksheet = writer.sheets['Tabelas_Registradas']
            worksheet.cell(row=1, column=1, value=f"Curso: {curso_nome}")
            start_row += 2
            for periodo, tabela in st.session_state['registros'][id_curso_selecionado].items():
                worksheet.cell(row=start_row, column=1, value=f"Período: {periodo}")
                start_row += 1
                tabela.to_excel(writer, sheet_name="Tabelas_Registradas", startrow=start_row, index=True)
                start_row += len(tabela) + 3
        output.seek(0)
        return output   

    def limpar_horario():
        st.session_state['agendas'] = {}
        st.session_state['registros'] = {}
        st.session_state['decisoes'] = {}

        # Remover o arquivo JSON específico do curso selecionado
        course_name = st.session_state['curso_selecionado']
        file_path = f'horarios/{course_name}_alocacoes.json'
        file_path2 = f'juncoes/{course_name}_juncao.json'

        if os.path.exists(file_path2):
            os.remove(file_path2)

        if os.path.exists(file_path):
            os.remove(file_path)
            st.warning(f"Horário do curso {course_name} removido.")

    def reset_confirmacao_exclusao():
        # Inicializar a flag no session_state se ainda não existir
        if 'reset_confirmacao_exclusao' not in st.session_state:
            st.session_state['reset_confirmacao_exclusao'] = False

        # Verificar se precisamos redefinir o checkbox
        if st.session_state['reset_confirmacao_exclusao']:
            st.session_state['confirmacao_exclusao'] = False
            st.session_state['reset_confirmacao_exclusao'] = False  

        # Criar a checkbox com a chave especificada
        st.checkbox(
            "Desejo limpar o horário e excluir o arquivo de alocação desse curso. CUIDADO: Não tem como recuperar o Horário excluído.",
            key='confirmacao_exclusao'
        )
        
        
            # Adicionar estilo CSS para as colunas da tabela
    
    def reset_confirmacao_upload():
        # Inicializar a flag no session_state se ainda não existir
        if 'reset_confirmacao_upload' not in st.session_state:
            st.session_state['reset_confirmacao_upload'] = False

        # Verificar se precisamos redefinir o checkbox
        if st.session_state['reset_confirmacao_upload']:
            st.session_state['confirmacao_upload'] = False
            st.session_state['reset_confirmacao_upload'] = False

        # Checkbox de confirmação
        st.checkbox(
            "Desejo carregar uma nova planilha de horário que substituirá o atual (CUIDADO: não tem como recuperar o horário antigo.)",
            key='confirmacao_upload'
        )

    def adicionar_css_customizado():
        st.markdown(
            """
            <style>
            .schedule-table {
                table-layout: fixed;
                width: 100%;
                word-wrap: break-word;
                border-collapse: collapse;
                font-size: 10px; 
            }
            .schedule-table th, .schedule-table td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
                vertical-align: top;
                width: 100px;
                white-space: pre-wrap;
                word-wrap: break-word;
            }
            .schedule-table th {
                background-color: #f2f2f2;
            }
            
            .schedule {
                background-color: #f2f2f2;
                font-size: 10px; 
                border-radius: 10px;
                text-align: center;
            }
            .conflito {
                background-color: #F08080;
                color: black;
            }
            .fusao {
                background-color: #87CEEB;
                color: black;
            }
    

            .custom-button {
                font-size: 16px; /* Tamanho do texto */
                padding: 8px 16px; /* Ajusta o tamanho do botão */
                background-color: #ff4d4d; /* Cor do botão */
                color: white; /* Cor do texto */
                border: none;
                border-radius: 5px;
                cursor: pointer;
                display: flex;
                align-items: center;
                justify-content: center;
            }
            """,
            unsafe_allow_html=True
        )

    def carrega_tabelas():
        # Carregar professores, cursos, períodos e horários de arquivos JSON
        professores_json = carregar_json('dados/professores.json')
        cursos_json = carregar_json('dados/cursos.json')
        periodos_json = carregar_json('dados/periodos.json')
        horarios_json = carregar_json('dados/horarios.json')
        dias_json = carregar_json('dados/dias.json')
        modalidades_json = carregar_json('dados/modalidade.json')

        # Converter JSON em DataFrames
        professores = pd.DataFrame(professores_json['professores'])
        cursos = pd.DataFrame(cursos_json['cursos'])
        periodos_df = pd.DataFrame(periodos_json['periodos'])
        horarios = horarios_json['horarios']
        dias = dias_json['dias']
        modalidades = pd.DataFrame(modalidades_json['modalidade'])

        return professores, cursos, periodos_df, horarios, dias, modalidades

    def inicializa_sessao():
        # Inicializar as agendas por curso e período no session_state
        if 'agendas' not in st.session_state:
            st.session_state['agendas'] = {}

        # Inicializar registro de tabelas para períodos
        if 'registros' not in st.session_state:
            st.session_state['registros'] = {}

        # Inicializar 'decisoes' no session_state se ainda não estiver presente
        if 'decisoes' not in st.session_state:
            st.session_state['decisoes'] = {}

        if 'conflitos' not in st.session_state:
            st.session_state['conflitos'] = {}

    def carrega_dados_na_alocacao(id_curso_selecionado):
        # Carregar os dados do JSON correspondente, se existir, após a troca de curso
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}
            dados_json = carregar_dados_alocacao(curso_selecionado)
            carregar_dados_no_session_state(dados_json, id_curso_selecionado)

    def incializa_agenda(id_curso_selecionado):
        # Inicializar a agenda para o curso e período selecionado, se não existir
        if id_curso_selecionado not in st.session_state['agendas']:
            st.session_state['agendas'][id_curso_selecionado] = {}

        if periodo_selecionado not in st.session_state['agendas'][id_curso_selecionado]:
            st.session_state['agendas'][id_curso_selecionado][periodo_selecionado] = create_schedule_table()

    def carrega_dados_tabelas():
        disciplinas = carregar_disciplinas(curso_selecionado)

        # Filtrar as disciplinas do curso selecionado
        disciplinas_opcoes = disciplinas[disciplinas['id_curso'] == id_curso_selecionado]['nome_disciplina'].tolist()
        professores_opcoes = professores.apply(lambda row: f"{row['matricula']} - {row['nome_professor']}", axis=1).tolist()
        modalidades_opcoes = modalidades['nome_modalidade'].tolist()
        return disciplinas, disciplinas_opcoes, professores_opcoes, modalidades_opcoes

    def carrega_planilha_horario(id_curso_selecionado):

        df = pd.read_excel(uploaded_file, sheet_name="Dados")
        df['Curso'] = df['Curso'].ffill()
        df['Horário'] = df['Horário'].ffill()
        df['Período'] = df['Período'].ffill()

        # Obter o nome do curso
        curso_nome = df['Curso'].iloc[0]

        # Inicializar estrutura de saída com o nome do curso
        output = {
            "curso": curso_nome,
            "periodos": {}
        }

        # Iterar sobre cada período, horário e dia da semana
        for period, period_df in df.groupby("Período"):
            period_list = []
            for horario, horario_df in period_df.groupby("Horário"):
                for day in ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]:
                    day_data = horario_df[day].dropna()
                    if not day_data.empty:
                        item = {
                            "horario": horario,
                            "dia": day,
                        }
                        for _, row in horario_df.iterrows():
                            info_key = row["Informação"].lower()
                            item[info_key] = row[day]

                        period_list.append(item)

            if period_list:
                output["periodos"][period] = period_list


        st.session_state['registros'][id_curso_selecionado] = {}
        st.session_state['decisoes'] = {}

        file_path = f'juncoes/{id_curso_selecionado}_juncao.json'
        if os.path.exists(file_path): os.remove(file_path)


        incializa_agenda(id_curso_selecionado) 
        if id_curso_selecionado not in st.session_state['registros']:
            st.session_state['registros'][id_curso_selecionado] = {}
        carregar_dados_no_session_state(output, id_curso_selecionado)

        with open(f'horarios/{curso_selecionado}_alocacoes.json', 'w', encoding='utf-8') as json_file:
            json.dump(output, json_file, indent=4)


        #return True
        #conflitos_curso_selecionado = verificar_conflitos_com_curso_selecionado("horarios", curso_selecionado)
        #if conflitos_curso_selecionado:
        #    registrar_conflitos(conflitos_curso_selecionado)
        #    salvar_juncoes(curso_selecionado)

    def transformar_conflitos(conflitos):
        resultado = []
        for conflito in conflitos:
            dia = conflito["dia"]
            horario = conflito["horario"]
            matricula = conflito["matricula_professor"]
            
            # Se houver mais de um período na lista de cursos, há um conflito
            if len(conflito["cursos"]) > 1:
                # Extrair apenas os períodos dos cursos
                periodos = [curso[1] for curso in conflito["cursos"]]
                # Adicionar ao resultado
                resultado.append([dia, horario, matricula] + periodos)
        
        return resultado

    def exibir_tabelas_periodo():
        if id_curso_selecionado in st.session_state['registros']:
            for periodo in periodos_df['nome_periodo']:
                if periodo in st.session_state['registros'][id_curso_selecionado]:
                    tabela = st.session_state['registros'][id_curso_selecionado][periodo].copy()

                    # Filtrar para exibir apenas "disciplina", "professor" e "modalidade" em cada célula
                    for time in tabela.index:
                        for day in tabela.columns:
                            cell_content = tabela.loc[time, day]
                            if pd.notna(cell_content) and cell_content:
                                # Extrair apenas disciplina e nome do professor
                                parts = cell_content.split(" - ")
                                if len(parts) >= 2:
                                    disciplina = parts[0]
                                    professor = parts[2]
                                    modalidade = parts[3]
                                    tabela.loc[time, day] = f"{disciplina} </br> {professor} </br> {modalidade}"

                    # Adicionar estilo de conflito ou fusão nas células com choque de horário
                    conflitos_detalhes = transformar_conflitos(conflitos_curso_selecionado)

                    for i, conflito in enumerate(conflitos_detalhes):
                        day, time, matricula_professor, *periodos = conflito  # Captura todos os períodos em conflito
                        conflito_key = f"{matricula_professor}_{day}_{time}"
                    
                        # Verifica se a decisão está armazenada no session_state e obtém apenas o tipo de decisão (Conflito ou Fusão)
                        tipo_conflito = st.session_state['decisoes'].get(conflito_key, {}).get("Tipo", "Conflito")
                        
                        # Define a classe de estilo com base no tipo de conflito
                        classe_estilo = "fusao" if tipo_conflito == "Fusão" else "conflito"

                        # Aplicar o estilo somente se a condição for satisfeita
                        if periodo in periodos and time in tabela.index and day in tabela.columns:
                            tabela.loc[time, day] = f"<div class='{classe_estilo}'>{tabela.loc[time, day]}</div>"


                    # Exibir a tabela estilizada
                    st.write(f"**Período: {periodo}**")
                    st.write(tabela.to_html(escape=False, classes="schedule-table"), unsafe_allow_html=True)

    def renderizar_agenda(id_curso_selecionado, periodo_selecionado):
        cols = st.columns(len(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns) + 1)
        cols[0].write("**Horários**")
        for i, day in enumerate(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns):
            cols[i + 1].write(f"**{day}**")

        for time_slot in horarios:
            cols = st.columns(len(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns) + 1)
            cols[0].write(f"**{time_slot}**")
            for i, day in enumerate(st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].columns):
                cell_content = st.session_state['agendas'][id_curso_selecionado][periodo_selecionado].loc[time_slot, day]
                if cell_content:
                    
                    parts = cell_content.split(" - ")
                    disciplina = parts[0]
                    matricula = parts[1]
                    professor = parts[2]
                    modalidade = parts[3]
                    data = f"{disciplina} </br> {professor} ({matricula}) </br> {modalidade}"
                    
                    cols[i + 1].markdown(f"<div class='schedule'>{data}</div>", unsafe_allow_html=True)
                    if cols[i + 1].button("🗑️",  key=f"desalocar_{time_slot}_{day}"):
                        deallocate_schedule(id_curso_selecionado, periodo_selecionado, day, time_slot)
                        st.rerun()
                else:
                    cols[i + 1].markdown("<div class='schedule'>-</div>", unsafe_allow_html=True)

    # Função para gerar PDF das tabelas registradas de cada período
    def gerar_pdf_tabelas_periodo(curso_selecionado):

        # Criar um buffer de memória para o PDF
        pdf_buffer = io.BytesIO()
        
        # HTML inicial para o PDF
        html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif; 
                    font-size: 8px; 
                }}
                .schedule-table {{ 
                    width: 100%; 
                    border-collapse: collapse; 
                }}
                .schedule-table th, .schedule-table td {{ 
                    border: 1px solid #ddd; 
                    padding: 8px; 
                    text-align: left; 
                    width: 100px; 
                    word-wrap: break-word; 
                    white-space: pre-wrap; 
                }}
                .schedule-table th {{ background-color: #f2f2f2; }}
                .conflito {{ background-color: #F08080; color: black; }}
                .fusao {{ background-color: #87CEEB; color: black; }}
                 /* Estilização para as colunas de Sexta, Sábado e Domingo */
                .schedule-table th:nth-child(6), /* Sexta */
                .schedule-table th:nth-child(7), /* Sábado */
                .schedule-table th:nth-child(8), /* Domingo */
                .schedule-table td:nth-child(6),
                .schedule-table td:nth-child(7),
                .schedule-table td:nth-child(8) {{
                    background-color: #F5F5DC;
                }}
            </style>
        </head>
        <body>
        <h2>Horário de {curso_selecionado}</h2>
        """

        # Adicionar o conteúdo HTML de cada tabela de período
        if id_curso_selecionado in st.session_state['registros']:
            for periodo in periodos_df['nome_periodo']:
                if periodo in st.session_state['registros'][id_curso_selecionado]:
                    tabela = st.session_state['registros'][id_curso_selecionado][periodo].copy()

                    # Estilizar a tabela de acordo com a função `exibir_tabelas_periodo`
                    for time in tabela.index:
                        for day in tabela.columns:
                            cell_content = tabela.loc[time, day]
                            if pd.notna(cell_content) and cell_content:
                                parts = cell_content.split(" - ")
                                if len(parts) >= 2:
                                    disciplina = parts[0]
                                    professor = parts[2]
                                    modalidade = parts[3]
                                    tabela.loc[time, day] = f"{disciplina} </br> {professor} </br> {modalidade}"

                    # Adicionar a tabela ao HTML
                    html += f"<h3>Período: {periodo}</h3>"
                    html += tabela.to_html(escape=False, classes="schedule-table")

    
        options = {
            "orientation": "Landscape"
        }
        # Gerar o PDF usando pdfkit e o HTML construído
        with tempfile.NamedTemporaryFile(delete=True, suffix=".pdf") as tmp_pdf:
            pdfkit.from_string(html, tmp_pdf.name, options=options)
            tmp_pdf.seek(0)  # Retornar para o início do arquivo temporário
            pdf_buffer.write(tmp_pdf.read())  # Lê o conteúdo para o buffer

        pdf_buffer.seek(0)  # Volta para o início do buffer
        return pdf_buffer      




    def verificar_conflito_interno(course_id, curso_selecionado, day, time, professor):
        matricula_professor = professor.split(' - ')[0]
        conflitos = []
        for periodo, agenda in st.session_state['registros'][course_id].items():
            cell_content = agenda.loc[time, day]
            if cell_content and matricula_professor in cell_content:
                conflitos.append(curso_selecionado)
        return conflitos

    def verificar_conflito_externo(diretorio_horarios, curso_selecionado, day, time, professor):
        matricula_professor = professor.split(' - ')[0]
        file = f'{curso_selecionado}_alocacoes.json'
        conflitos = []
        for arquivo in os.listdir(diretorio_horarios):
            if arquivo.endswith('_alocacoes.json') and arquivo != file:
                with open(os.path.join(diretorio_horarios, arquivo), 'r', encoding='utf-8') as json_file:
                    dados = json.load(json_file)
                    curso = dados["curso"]
                    for periodo, alocacoes in dados["periodos"].items():
                        for alocacao in alocacoes:    
                            if alocacao["dia"] == day and alocacao["horario"] == time and alocacao["matricula_professor"] == matricula_professor:
                                conflitos.append(curso)

        return conflitos

    def registrar_conflito(conflito_key, tipo, cursos, curso_selecionado, periodo_selecionado, day, time, professor):
        matricula_professor = professor.split(' - ')[0]
        nome_professor  = professor.split(' - ')[1]

        st.session_state['conflitos'][conflito_key] = {
            "Tipo": tipo,
            "Curso": curso_selecionado,
            "Período": periodo_selecionado,
            "Dia": day,
            "Horário": time,
            "Matrícula": matricula_professor,
            "Professor": nome_professor 
        }


    def salvar_conflito_multicurso(conflito_key, tipo, cursos, curso_selecionado, periodo_selecionado, day, time, professor):
        # Extrair matrícula e nome do professor a partir do campo `professor`
        matricula_professor = professor.split(' - ')[0]
        nome_professor = professor.split(' - ')[1]

        # Estrutura do conflito que será salva
        conflito = {
            conflito_key: {
                "Tipo": tipo,
                "Curso": curso_selecionado,
                "Período": periodo_selecionado,
                "Dia": day,
                "Horário": time,
                "Matrícula": matricula_professor,
                "Professor": nome_professor
            }
        }

        pasta_juncao = "juncoes"
        nome_arquivo_selecionado = f"{curso_selecionado}_juncao.json"
        caminho_arquivo = os.path.join(pasta_juncao, nome_arquivo_selecionado)
        with open(caminho_arquivo, "w", encoding='utf-8') as f:
            dados = st.session_state['conflitos'].values()
            json.dump(list(dados), f, ensure_ascii=False, indent=4)

        st.write(cursos)
        # Iterar sobre cada curso envolvido e atualizar os conflitos
        for nome_curso in cursos:
            nome_arquivo = f"{nome_curso}_juncao.json"
            caminho_arquivo = os.path.join(pasta_juncao, nome_arquivo)
        
            if nome_arquivo_selecionado != nome_arquivo:
                
                # Carregar os conflitos existentes para o curso se o arquivo já existir
                if os.path.exists(caminho_arquivo):
                    with open(caminho_arquivo, "r", encoding='utf-8') as f:
                        conflitos_existentes = json.load(f)
                else:
                    conflitos_existentes = {}

                conflito_data = conflito[conflito_key]
                st.write(caminho_arquivo)
    
                conflito_existe = any(
                    c.get("Dia") == conflito_data["Dia"] and
                    c.get("Horário") == conflito_data["Horário"] and
                    c.get("Matrícula") == conflito_data["Matrícula"] and
                    c.get("Professor") == conflito_data["Professor"]
                    for c in conflitos_existentes
                )
               
                if conflito_existe:
                    conflitos_existentes.append(conflito_data)
                    st.write(conflito_data)
                    # Salvar o arquivo atualizado de conflitos para o curso
                    with open(caminho_arquivo, "w", encoding='utf-8') as f:
                        json.dump(conflitos_existentes, f, ensure_ascii=False, indent=4)

    def carregar_conflitos(curso_selecionado):
        if 'conflitos' not in st.session_state:
            st.session_state['conflitos'] = {}

        pasta_juncao = "juncoes"
        caminho_arquivo = os.path.join(pasta_juncao, f"{curso_selecionado}_juncao.json")
        
        if os.path.exists(caminho_arquivo):
            with open(caminho_arquivo, "r", encoding='utf-8') as f:
                conflitos = json.load(f)
                # Verificar se o conteúdo é um dicionário no formato esperado
                if isinstance(conflitos, dict) and all(isinstance(value, dict) for value in conflitos.values()):
                    st.session_state['conflitos'] = conflitos
                else:
                    st.session_state['conflitos'] = {}  # Reiniciar se o formato for inesperado
                    st.warning("O arquivo de conflitos não está no formato esperado.")
        else:
            st.session_state['conflitos'] = {}  # Reiniciar se o arquivo não existir

       


#####################################################################################
    st.write("### Montagem de Horário por Curso")

    salva_conflito = False
    professores, cursos, periodos_df, horarios, dias, modalidades = carrega_tabelas()
    adicionar_css_customizado()
    inicializa_sessao()


    # Combobox para a escolha do curso
    col1, col2 = st.columns(2)
    with col1:
        curso_selecionado = st.selectbox("Selecione o curso:", cursos['nome_curso'], key="curso_selecionado", on_change=limpar_secao_ao_trocar_curso)
        carregar_conflitos(curso_selecionado)
        reset_confirmacao_upload()
        uploaded_file = st.file_uploader("Carregar Planilha de Horário", type="xlsx") if st.session_state.get('confirmacao_upload') else None
    with col2: periodo_selecionado = st.selectbox("Selecione o período:", periodos_df['nome_periodo'].tolist())
    id_curso_selecionado = cursos[cursos['nome_curso'] == curso_selecionado]['id_curso'].values[0]


    carrega_dados_na_alocacao(id_curso_selecionado)    
    incializa_agenda(id_curso_selecionado)     
    disciplinas, disciplinas_opcoes, professores_opcoes, modalidades_opcoes = carrega_dados_tabelas()    


    # Leitura de planilha
    if uploaded_file and st.session_state.get('confirmacao_upload'):
        #salva_conflito = carrega_planilha_horario(id_curso_selecionado)
        carrega_planilha_horario(id_curso_selecionado)
        st.session_state['reset_confirmacao_upload'] = True     
        st.success("Planilha carregada com sucesso.")


    # Combobox para as escolhas de "Dia da Semana" e "Horário" 
    st.write("### Alocação de Disciplina e Professor")
    col3, col4 = st.columns(2)
    with col3: day = st.selectbox("Dia da Semana", dias)
    with col4: time = st.selectbox("Horário", horarios)


    # Combobox para as escolhas de "Disciplina" e "Professor"
    col5, col6 = st.columns(2)
    with col5:  discipline = st.selectbox("Disciplina", disciplinas_opcoes)
    with col6: modalidade = st.selectbox("Modalidade", modalidades_opcoes)
        

    # Alocar disciplina e professor# Criar uma nova linha para os dois botões
    col7, col8 = st.columns(2)
    with col7:
        professor = st.selectbox("Professor", professores_opcoes)
        alocar_button = st.button("Alocar Disciplina e Professor", key="alocar", use_container_width=True)
    if alocar_button:
        allocate_schedule(id_curso_selecionado, periodo_selecionado, day, time, discipline, professor, modalidade)
    

    # Renderizar a agenda de definição de horário
    st.write(f"### Programação de Horários - {periodo_selecionado}")
    renderizar_agenda(id_curso_selecionado, periodo_selecionado)
    col_button, _ = st.columns([1, 2])  
    with col_button:
        registrar_button = st.button("Registrar Horário", key="registrar", use_container_width=True)         
    
    if registrar_button:
        registrar_horario()
        conflito_interno = verificar_conflito_interno(id_curso_selecionado, curso_selecionado, day, time, professor)
        conflito_externo = verificar_conflito_externo("horarios", curso_selecionado, day, time, professor)

        if conflito_interno or conflito_externo:
            st.error("Conflito detectado! Verifique e resolva os conflitos antes de prosseguir.")
            cursos_envolvidos = set(conflito_interno + conflito_externo)
            st.write(cursos_envolvidos)
            conflito_key = f"{professor}_{day}_{time}"
            registrar_conflito(conflito_key, "Conflito", cursos_envolvidos, curso_selecionado, periodo_selecionado, day, time, professor)
            salvar_conflito_multicurso(conflito_key, "Conflito", cursos_envolvidos, curso_selecionado, periodo_selecionado, day, time, professor)






    conflitos_curso_selecionado = verificar_conflitos_com_curso_selecionado("horarios", curso_selecionado)
   
    # Exibir as tabelas registradas de cada período
    st.write("### Horários Definidos por Período")
    exibir_tabelas_periodo()

   # Botão para gerar e baixar o PDF
    if st.button("Baixar PDF do Horário"):
        pdf_file = gerar_pdf_tabelas_periodo(curso_selecionado)
        st.download_button(
            label="Download PDF",
            data=pdf_file,
            file_name="horarios_geral.pdf",
            mime="application/pdf"
        )


    reset_confirmacao_exclusao()   
    if st.session_state.get('confirmacao_exclusao'):
        if st.button("Limpar Horário e Recomeçar"):
            limpar_horario()
            st.success("Horário e arquivo de alocação excluídos com sucesso.")
            st.session_state['reset_confirmacao_exclusao'] = True 
            st.rerun()
    else:
        st.warning("Por favor, marque a confirmação para visualizar o botão de exclusão.")


    st.write("\n")
    if st.button("Gerar Excel"):
        excel_file = gerar_excel()
        st.download_button(label="Download Excel", data=excel_file, file_name="tabelas_registradas.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
         